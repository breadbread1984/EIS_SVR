#!/usr/bin/python3

from absl import flags, app
from shutil import rmtree
from os import mkdir
from os.path import join, exists
from sklearn.svm import SVR
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler
import pickle
from openpyxl import load_workbook
import numpy as np

FLAGS = flags.FLAGS

def add_options():
  flags.DEFINE_string('input_dir', default = None, help = 'path to dataset')
  flags.DEFINE_string('ckpt', default = 'ckpt', help = 'path to ckpt')
  flags.DEFINE_boolean('test', default = False, help = 'whether to run in test mode')

def load_dataset():
  voltage = load_workbook(join(FLAGS.input_dir, 'Voltage.xlsx')); voltage = voltage.active
  current = load_workbook(join(FLAGS.input_dir, 'Current.xlsx')); current = current.active
  real = load_workbook(join(FLAGS.input_dir, 'EIS_real.xlsx')); real = real.active
  imag = load_workbook(join(FLAGS.input_dir, 'EIS_imag.xlsx')); imag = imag.active
  samples = list()
  for col in range(voltage.max_column):
    v, c = list(), list()
    for row in range(voltage.max_row):
      v.append(voltage.cell(row = row + 1, column = col + 1).value)
      c.append(current.cell(row = row + 1, column = col + 1).value)
    v = np.array(v)
    c = np.array(c)
    pulse = np.stack([v,c], axis = -1) # pulse.shape = (length, 2)
    r, i = list(), list()
    for row in range(real.max_row):
      r.append(real.cell(row = row + 1, column = col + 1).value)
      i.append(imag.cell(row = row + 1, column = col + 1).value)
    r = np.array(r)
    i = np.array(i)
    eis = np.stack([r,i], axis = -1) # eis.shape = (length, 2)
    samples.append((pulse, eis))
  return samples

def main(unused_argv):
  if FLAGS.test: test()
  else: train()

def test():
  models = list()
  for i in range(35 * 2):
    with open('%d.pickle' % i, 'rb') as f:
      models.append(pickle.loads(f.read()))
  samples = load_dataset()
  x = np.stack([sample[0].flatten() for sample in samples], axis = 0) # x.shape = (sample_num, 1800*2)
  y = np.stack([sample[1].flatten() for sample in samples], axis = 0) # y.shape = (sample_num, 35*2)
  losses = list()
  for idx, (sample, label) in enumerate(zip(x, y)):
    results = list()
    for model in models:
      result = model.predict(np.expand_dims(sample, axis = 0)) # y.shape = (1, 1)
      results.append(result)
    results = np.squeeze(np.concat(results, axis = -1), axis = 0) # results.shape = (1, 35 * 2)
    mae = np.mean(np.abs(label - results))
    print('#%d mea: %f' % (idx, mae))
    losses.append(mae)
  print('mean mae: %f' % np.mean(losses))

def train():
  samples = load_dataset()
  if exists(FLAGS.ckpt): rmtree(FLAGS.ckpt)
  mkdir(FLAGS.ckpt)
  x = np.stack([sample[0].flatten() for sample in samples], axis = 0) # x.shape = (sample_num, 1800*2)
  y = np.stack([sample[1].flatten() for sample in samples], axis = 0) # y.shape = (sample_num, 35*2)
  models = [SVR(C=1.0, epsilon = 0.2) for i in range(35*2)]
  for i in range(35 * 2):
    regr = make_pipeline(StandardScaler(), models[i])
    regr.fit(x, y[:,i])
    with open(join(FLAGS.ckpt, '%d.pickle' % i), 'wb') as f:
      f.write(pickle.dumps(models[i]))

if __name__ == "__main__":
  add_options()
  app.run(main)

